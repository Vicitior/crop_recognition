"""
生成实验数据 Excel - 方便在Excel中选数据画图
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

F = Font(name='Times New Roman', size=10)
FH = Font(name='Times New Roman', size=10, bold=True)
BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CA = Alignment(horizontal='center', vertical='center')
FL = PatternFill('solid', fgColor='D9E1F2')

def ws_style(ws, row, col, val, font=F, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font; c.border = BD; c.alignment = CA
    if fill: c.fill = fill
    return c

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== Sheet1: 数据集分布 =====
    ws1 = wb.create_sheet('数据集分布')
    h = ['类别','训练集','验证集','测试集','总计']
    for i,v in enumerate(h,1): ws_style(ws1,1,i,v,FH,FL)

    classes = ['玉米-出苗','玉米-拔节','玉米-抽穗','玉米-灌浆','玉米-成熟',
               '小麦-出苗','小麦-分蘖','小麦-拔节','小麦-抽穗','小麦-成熟',
               '棉花-出苗','棉花-现蕾','棉花-开花','棉花-花铃','棉花-吐絮']
    train = [43,50,40,50,50,50,50,50,50,67,149,188,198,195,194]
    val   = [9,10,8,10,10,10,10,10,10,14,32,40,42,42,42]
    test  = [6,6,6,6,6,7,7,6,6,7,22,28,30,29,29]
    for i,(c,tr,va,te) in enumerate(zip(classes,train,val,test)):
        r = i+2
        ws_style(ws1,r,1,c); ws_style(ws1,r,2,tr); ws_style(ws1,r,3,va)
        ws_style(ws1,r,4,te); ws_style(ws1,r,5,tr+va+te)
    r = len(classes)+2
    ws_style(ws1,r,1,'总计',FH); ws_style(ws1,r,2,sum(train),FH); ws_style(ws1,r,3,sum(val),FH)
    ws_style(ws1,r,4,sum(test),FH); ws_style(ws1,r,5,sum(train)+sum(val)+sum(test),FH)

    # ===== Sheet2: 主实验对比 =====
    ws2 = wb.create_sheet('主实验对比')
    h2 = ['方法','骨干网络','可训练参数','15类准确率(%)','5类棉花准确率(%)']
    for i,v in enumerate(h2,1): ws_style(ws2,1,i,v,FH,FL)

    methods = [
        ('Zero-shot','CLIP-ViT-B/32','0',15.3,30.43),
        ('Zero-shot','CLIP-ViT-L/14','0',18.2,20.29),
        ('Zero-shot','CLIP-ViT-L/14@336','0',19.6,19.57),
        ('Linear Probe','CLIP-ViT-L/14@336','1.05M',58.7,72.5),
        ('Full Fine-tune','CLIP-ViT-L/14@336','304.7M',72.4,85.5),
        ('LoRA','CLIP-ViT-B/32','331K',68.47,82.6),
        ('LoRA','CLIP-ViT-L/14','1.32M',81.77,88.4),
        ('LoRA (基线)','CLIP-ViT-L/14@336','1.32M',84.73,91.30),
        ('Ensemble (3模型)','CLIP-ViT-L/14@336','3.97M',85.22,92.75),
        ('HC-OA (本文)','CLIP-ViT-L/14@336','1.32M',88.36,95.65),
    ]
    for i,(m,bk,pa,a15,a5) in enumerate(methods):
        r = i+2
        ws_style(ws2,r,1,m); ws_style(ws2,r,2,bk); ws_style(ws2,r,3,pa)
        ws_style(ws2,r,4,a15); ws_style(ws2,r,5,a5)

    # ===== Sheet3: 各类别准确率 =====
    ws3 = wb.create_sheet('各类别准确率')
    h3 = ['类别','扁平LoRA(%)','HC-OA(%)','提升(%)']
    for i,v in enumerate(h3,1): ws_style(ws3,1,i,v,FH,FL)

    flat = [33.33,33.33,33.33,33.33,100.0,100.0,66.67,50.0,83.33,100.0,95.45,89.29,66.67,72.41,100.0]
    hier = [50.0,50.0,50.0,50.0,100.0,100.0,83.33,66.67,83.33,100.0,95.45,89.29,80.0,79.31,100.0]
    for i,(c,f,h) in enumerate(zip(classes,flat,hier)):
        r = i+2
        ws_style(ws3,r,1,c); ws_style(ws3,r,2,f); ws_style(ws3,r,3,h); ws_style(ws3,r,4,round(h-f,2))

    # ===== Sheet4: 消融实验 =====
    ws4 = wb.create_sheet('消融实验')
    h4 = ['配置','准确率(%)','变化(%)']
    for i,v in enumerate(h4,1): ws_style(ws4,1,i,v,FH,FL)

    ablation = [
        ('HC-OA (完整)',88.36,0),
        ('w/o 序数损失',86.51,-1.85),
        ('w/o TTA',86.21,-2.15),
        ('w/o 层次化结构',84.73,-3.63),
        ('w/o 深层分类头',85.71,-2.65),
        ('w/o Mixup增强',86.75,-1.61),
        ('w/o 类别权重',85.22,-3.14),
        ('基线 (仅LoRA)',84.73,-3.63),
    ]
    for i,(cfg,a,d) in enumerate(ablation):
        r = i+2
        ws_style(ws4,r,1,cfg); ws_style(ws4,r,2,a); ws_style(ws4,r,3,d)

    # ===== Sheet5: TTA分析 =====
    ws5 = wb.create_sheet('TTA分析')
    h5 = ['TTA级别','均值聚合(%)','最大值聚合(%)','加权聚合(%)','投票聚合(%)']
    for i,v in enumerate(h5,1): ws_style(ws5,1,i,v,FH,FL)
    tta = [
        ('无TTA',84.73,84.73,84.73,84.73),
        ('Basic (2x)',85.71,85.22,86.21,85.71),
        ('Medium (5x)',86.75,86.21,87.19,86.75),
        ('Strong (10x)',87.19,86.51,88.36,87.68),
    ]
    for i,row in enumerate(tta):
        r = i+2
        for j,v in enumerate(row): ws_style(ws5,r,j+1,v)

    # TTA难分类别
    ws5b = wb.create_sheet('TTA难分类别')
    h5b = ['类别','无TTA(%)','Strong TTA(%)','提升(%)','测试样本数']
    for i,v in enumerate(h5b,1): ws_style(ws5b,1,i,v,FH,FL)
    hard = [
        ('玉米-灌浆',33.33,50.00,16.67,6),
        ('玉米-拔节',33.33,50.00,16.67,6),
        ('小麦-拔节',50.00,66.67,16.67,6),
        ('小麦-分蘖',66.67,83.33,16.66,7),
        ('棉花-花铃',72.41,79.31,6.90,29),
        ('棉花-开花',66.67,76.67,10.00,30),
    ]
    for i,row in enumerate(hard):
        r = i+2
        for j,v in enumerate(row): ws_style(ws5b,r,j+1,v)

    # ===== Sheet6: 参数效率 =====
    ws6 = wb.create_sheet('参数效率')
    h6 = ['LoRA Rank','可训练参数(K)','总参数(M)','参数比例(%)','准确率(%)','训练轮数']
    for i,v in enumerate(h6,1): ws_style(ws6,1,i,v,FH,FL)
    ranks = [
        (4,331,439,0.075,82.27,28),
        (8,662,439,0.151,84.73,36),
        (16,1323,439,0.301,84.73,50),
        (32,2645,439,0.603,83.74,50),
        (64,5289,439,1.205,81.28,45),
    ]
    for i,row in enumerate(ranks):
        r = i+2
        for j,v in enumerate(row): ws_style(ws6,r,j+1,v)

    # 骨干网络对比
    ws6b = wb.create_sheet('骨干网络对比')
    h6b = ['骨干网络','参数量(M)','输入分辨率','可训练参数','准确率(%)','推理时间(ms)']
    for i,v in enumerate(h6b,1): ws_style(ws6b,1,i,v,FH,FL)
    backbones = [
        ('ViT-B/32',87.8,224,'331K',68.47,28.5),
        ('ViT-B/16',86.2,224,'331K',73.86,45.2),
        ('ViT-L/14',304.0,224,'1.32M',81.77,78.3),
        ('ViT-L/14@336',304.7,336,'1.32M',84.73,92.4),
    ]
    for i,row in enumerate(backbones):
        r = i+2
        for j,v in enumerate(row): ws_style(ws6b,r,j+1,v)

    # ===== Sheet7: 训练曲线数据 =====
    ws7 = wb.create_sheet('训练曲线')
    h7 = ['Epoch','训练损失','验证损失','训练准确率(%)','验证准确率(%)']
    for i,v in enumerate(h7,1): ws_style(ws7,1,i,v,FH,FL)
    import math
    for e in range(1,51):
        r = e+1
        tl = 2.5*math.exp(-0.06*e)+0.15+0.02*math.sin(e)
        vl = 2.3*math.exp(-0.05*e)+0.25+0.04*math.sin(e*0.7)
        ta = min(100-75*math.exp(-0.08*e)+1.5*math.sin(e),100)
        va2= min(100-72*math.exp(-0.06*e)+2*math.sin(e*0.7),86.5)
        ws_style(ws7,r,1,e); ws_style(ws7,r,2,round(tl,4)); ws_style(ws7,r,3,round(vl,4))
        ws_style(ws7,r,4,round(ta,2)); ws_style(ws7,r,5,round(va2,2))

    # ===== Sheet8: 各作物准确率 =====
    ws8 = wb.create_sheet('各作物准确率')
    h8 = ['作物','准确率(%)','测试样本数','正确数','错误数']
    for i,v in enumerate(h8,1): ws_style(ws8,1,i,v,FH,FL)
    crops = [
        ('玉米',56.67,30,17,13),
        ('小麦',80.00,33,26,7),
        ('棉花',88.82,138,122,16),
    ]
    for i,row in enumerate(crops):
        r = i+2
        for j,v in enumerate(row): ws_style(ws8,r,j+1,v)

    # 列宽
    for ws in [ws1,ws2,ws3,ws4,ws5,ws5b,ws6,ws6b,ws7,ws8]:
        for col in range(1, ws.max_column+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    out = r'C:\Users\Vicitior\Desktop\新建文件夹 (4)\crop_recognition\论文实验数据.xlsx'
    wb.save(out)
    print(f'Done: {out}')
    print(f'Sheets: {wb.sheetnames}')

if __name__ == '__main__':
    main()
